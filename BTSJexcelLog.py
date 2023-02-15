import Levenshtein
import glob
import MeCab
import openpyxl
import pandas as pd
import math
import re

from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.borders import Border, Side

# システム発話の文末判定用関数
def leven(str1, str2):
    lev_dist = Levenshtein.distance(str1, str2)
    devider = len(str1) if len(str1) > len(str2) else len(str2)
    lev_dist = lev_dist / devider
    lev_dist = 1 - lev_dist

    return lev_dist

# 文末判定リストを用いた文末判定用関数
def listFind(list, string):
	for l in list:
		if string.find(l) >= 0:
			return True
	
	return False

# 発話文の分割用関数
def listSlice(int):
    # 発話文の開始と終了のindex格納
    eosIListS = []
    eosIListE = []

    # 1文
    if int == 1:
        startS = 0
        eosIListS.append(startS)
        eosIListE.append(len(dialogueLog[i][2]) - 1)
    # 2文以上
    else:
        startS = 0

        # 発話文毎ループ
        for l in range(int):
            eosIP = dialogueLog[i][2].find("。", startS)
            eosIE = dialogueLog[i][2].find("！", startS)
            eosIQ = dialogueLog[i][2].find("？", startS)

            eosIListS.append(startS)

            # 3種類
            if eosIP >= 0 and eosIE >= 0 and eosIQ >= 0 and eosIP < eosIE and eosIP < eosIQ:
                eosIListE.append(eosIP)
            elif eosIP >= 0 and eosIE >= 0 and eosIQ >= 0 and eosIE < eosIP and eosIE < eosIQ:
                eosIListE.append(eosIE)
            elif eosIP >= 0 and eosIE >= 0 and eosIQ >= 0 and eosIQ < eosIP and eosIQ < eosIE:
                eosIListE.append(eosIQ)
            # 2種類
            elif eosIP >= 0 and eosIE >= 0 and eosIP < eosIE:
                eosIListE.append(eosIP)
            elif eosIP >= 0 and eosIE >= 0 and eosIE < eosIP:
                eosIListE.append(eosIE)
            elif eosIP >= 0 and eosIQ >= 0 and eosIP < eosIQ:
                eosIListE.append(eosIP)
            elif eosIP >= 0 and eosIQ >= 0 and eosIQ < eosIP:
                eosIListE.append(eosIQ)
            elif eosIE >= 0 and eosIQ >= 0 and eosIE < eosIQ:
                eosIListE.append(eosIE)
            elif eosIE >= 0 and eosIQ >= 0 and eosIQ < eosIE:
                eosIListE.append(eosIQ)
            # 1種類
            elif eosIP>= 0:
                eosIListE.append(eosIP)
            elif eosIE>= 0:
                eosIListE.append(eosIE)
            elif eosIQ>= 0:
                eosIListE.append(eosIQ)
            # なし
            else:
                eosIListE.append(len(dialogueLog[i][2]) - 1)

            startS = eosIListE[l] + 1

    return eosIListS, eosIListE
    
# ファイルパス一覧取得
fileList = glob.glob("./dialogueLog/**/*.xlsx")

# MeCabの事前設定
wakati = MeCab.Tagger("-Owakati")
wakati.parse("")
chasen = MeCab.Tagger("-Ochasen")
chasen.parse("")

# 各条件の判定カウント用変数の定義
teineitaiUserCountP = 0
teineitaiUserCountP1 = 0
teineitaiUserCountP2 = 0
teineitaiUserCountNM = 0
teineitaiUserCountN = 0
hiteineitaiUserCountP = 0
hiteineitaiUserCountP1 = 0
hiteineitaiUserCountP2 = 0
hiteineitaiUserCountNM = 0
hiteineitaiUserCountN = 0
switchUserCountP = 0
switchUserCountP1 = 0
switchUserCountP2 = 0
switchUserCountNM = 0
switchUserCountN = 0

# 判定用リスト
teikeiList1 = ["こんにちは。", "ごめんなさい。", "はじめまして。", "こんばんは。", "初めまして。"]
teikeiList2 = ["こんにちは！", "ごめんなさい！", "はじめまして！", "こんばんは！", "初めまして！", "こんにちは？", "ごめんなさい？", "はじめまして？", "こんばんは？", "初めまして？"]
NMlist = ["うん","うん。","うん！","うん？","そう", "そう。", "そう！", "そう？", "はい", "はい。", "はい！", "はい？", "あ", "あ。", "あ！", "あ？", "え", "え。", "え！", "え？", 
          "うーん", "うーん。", "うーん！", "うーん？", "ね", "ね。", "ね！", "ね？", "どうぞ", "どうぞ。", "どうぞ！", "どうぞ？", "まあね", "まあね。", "まあね！", "まあね？", 
          "よろ", "よろ。", "よろ！", "よろ？", "おう", "おう。", "おう！", "おう？", "ありあり", "ありあり。", "ありあり！", "ありあり？", "どうも", "どうも。", "どうも！", "どうも？", 
          "いえいえ", "いえいえ。", "いえいえ！", "いえいえ？", "もちろん", "もちろん。", "もちろん！", "もちろん？", "あー", "あー。", "あー！", "あー？", "えー", "えー。", "えー！", "えー？", 
          "いやいや", "いやいや。", "いやいや！", "いやいや？", "そだねー", "そだねー。", "そだねー！", "そだねー？", "そうだねぇ", "そうだねぇ。", "そうだねぇ！", "そうだねぇ？", 
          "そうだね", "そうだね。", "そうだね！", "そうだね？", "そうだな", "そうだな。", "そうだな！", "そうだな？", "せやな", "せやな。", "せやな！", "せやな？", "はーい", "はーい。", "はーい！", "はーい？", 
          "ぜひ", "ぜひ。", "ぜひ！", "ぜひ？", "まあ", "まあ。", "まあ！", "まあ？", "いえ", "いえ。", "いえ！", "いえ？", "はいよ", "はいよ。", "はいよ！", "はいよ？", "うい", "うい。", "うい！", "うい？", 
          "ううん", "ううん。", "ううん！", "ううん？", "そだね", "そだね。", "そだね！", "そだね？", "うむ", "うむ。", "うむ！", "うむ？", "そうね", "そうね。", "そうね！", "そうね？", 
          "うんうん", "うんうん。", "うんうん！", "うんうん？", "そうだねー", "そうだねー。", "そうだねー！", "そうだねー？", "そうだなー","そうだなー。","そうだなー！","そうだなー？",
          "いいえ", "いいえ。", "いいえ！", "いいえ？", "だなー", "だなー。", "だなー！", "だなー？", "だね", "だね。", "だね！", "だね？", "いや", "いや。", "いや！", "いや？", 
          "なるほど", "なるほど。", "なるほど！", "なるほど？", "わかる", "わかる。", "わかる！", "わかる？", "わかった", "わかった。", "わかった！", "わかった？", "わかったよ", "わかったよ。", "わかったよ！", "わかったよ？", 
          "いいよ", "いいよ。", "いいよ！", "いいよ？", "いいよー", "いいよー。", "いいよー！", "いいよー？", "いいよね", "いいよね。", "いいよね！", "いいよね？", "ええんやで", "ええんやで。", "ええんやで！", "ええんやで？", 
          "いいだろう", "いいだろう。", "いいだろう！", "いいだろう？", "なんだろう", "なんだろう。", "なんだろう！", "なんだろう？", "とんでもない", "とんでもない。", "とんでもない！", "とんでもない？", 
          "なる", "なる。", "なる！", "なる？", "そんなことないよ", "そんなことないよ。", "そんなことないよ！", "そんなことないよ？", "とくには", "とくには。", "とくには！", "とくには？", 
          "いいけど", "いいけど。", "いいけど！", "いいけど？", "まあーまあー", "まあーまあー。", "まあーまあー！", "まあーまあー？", "もちろん", "もちろん。", "もちろん！", "もちろん？", 
          "それほどでも", "それほどでも。", "それほどでも！", "それほどでも？", "あまり", "あまり。", "あまり！", "あまり？", "また", "また。", "また！", "また？", "特に", "特に。", "特に！", "特に？", 
          "うふふ", "うふふ。", "うふふ！", "うふふ？", "たまに", "たまに。", "たまに！", "たまに？", "そうかも", "そうかも。", "そうかも！", "そうかも？", "いますぐ", "いますぐ。", "いますぐ！", "いますぐ？"]

# ファイル毎ループ
for file in fileList:
    # xlsxの読み込み
    with open(file, "rb") as f:
        df = openpyxl.load_workbook(f)
        ds = df.active

    # 新規作成
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    # 項目名
    ws["A1"].value = "ターン番号"
    ws["B1"].value = "話者"
    ws["C1"].value = "発話内容"
    ws["D1"].value = "ユーザ文末"
    ws["E1"].value = "システム文末"
    ws["F1"].value = "ユーザスピーチレベルシフト"

    # 高さと幅
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 71
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 27

    # 背景色
    userbcColor = PatternFill(patternType="solid", fgColor="ccffff")
    systembcColor = PatternFill(patternType="solid", fgColor="ffff99")

    # 罫線パターン
    thin = Side(style='thin', color='000000')
    dotted = Side(style='dotted', color='000000')
    thindotBorder = Border(top=dotted, bottom=dotted, left=thin, right=thin)
    thinBorder = Border(top=thin, bottom=thin, left=thin, right=thin)
    topdotBorder = Border(top=dotted, bottom=thin, left=thin, right=thin)
    
    # 罫線
    ws["A1"].border = thinBorder
    ws["B1"].border = thinBorder
    ws["C1"].border = thinBorder
    ws["D1"].border = thinBorder
    ws["E1"].border = thinBorder
    ws["F1"].border = thinBorder

    # 前回の判定結果
    lastTimeEos = []

    # 今回の判定結果
    nowEos = []

    # ユーザ文末判定のカウント
    userCountP = 0
    userCountP1 = 0
    userCountP2 = 0
    userCountNM = 0
    userCountN = 0 

    # システム文末判定のカウント
    sysCountT = 0
    sysCountH = 0
    sysCountK = 0

    # 発話文
    dialogueLog = []

    # 発話文（dialogueLogをlist化する用）
    d_list = []

    # 行毎ループ
    for row in ds:
        dialogueLog.append([col.value for col in row])

    # ログのターン数毎にループ
    for i in range(1, len(dialogueLog)):
        # 各項目へ書き込み
        ws[i + 1][0].value = i - 1
        ws[i + 1][1].value = dialogueLog[i][1]
        ws[i + 1][2].value = dialogueLog[i][2]
        
        # ユーザ発話
        if dialogueLog[i][1] == "user":
            backColor = userbcColor

            # 発話内容が空白だったらログにnullを入れる
            if ws[i + 1][2].value is None:
                dialogueLog[i][2] = "null"

            # 。！？が連続してある時は1文字分として置換する
            while (re.search("。{2,}", dialogueLog[i][2]) != None):
                dialogueLog[i][2] = re.sub("。{2,}", "。", dialogueLog[i][2])
            while (re.search("！{2,}", dialogueLog[i][2]) != None):
                dialogueLog[i][2] = re.sub("！{2,}", "！", dialogueLog[i][2])
            while (re.search("？{2,}", dialogueLog[i][2]) != None):
                dialogueLog[i][2] = re.sub("？{2,}", "？", dialogueLog[i][2])

            # wakatiとchasenを行い，リスト化
            wakatiList = wakati.parse(dialogueLog[i][2]).split(" ")[:-1]
            chasenList = [x.split("\t") for x in chasen.parse(dialogueLog[i][2]).split("\n")]
            chasenTList = pd.DataFrame(chasenList).T.values.tolist()

            # 判定に不要な文字を削除する（本プログラムでは☺，笑，（笑））
            while dialogueLog[i][2].find("☺") >= 0 and not((len(dialogueLog[i][2]) == 2 and "記号" in chasenTList[3][-3]) or len(dialogueLog[i][2]) == 1 or (wakatiList.index("☺") >= 1 and "記号" in chasenTList[3][wakatiList.index("☺") - 1])):
                d_list = list(dialogueLog[i][2])
                d_del = d_list.index("☺")
                del d_list[d_del]
                dialogueLog[i][2] = "".join(d_list)
            while dialogueLog[i][2].find("笑") >= 0 and not((len(dialogueLog[i][2]) == 2 and "記号" in chasenTList[3][-3]) or len(dialogueLog[i][2]) == 1 or (wakatiList.index("笑") >= 1 and "記号" in chasenTList[3][wakatiList.index("笑") - 1])):
                d_list = list(dialogueLog[i][2])
                d_del = d_list.index("笑")
                del d_list[d_del]
                dialogueLog[i][2] = "".join(d_list)
            while dialogueLog[i][2].find("（笑）") >= 0:
                d_list = list(dialogueLog[i][2])
                d_del = dialogueLog[i][2].find("（笑）")
                del d_list[d_del:d_del + 3]
                dialogueLog[i][2] = "".join(d_list)

            # 再度wakatiとchasenを行い，リスト化
            wakatiList = wakati.parse(dialogueLog[i][2]).split(" ")[:-1]
            chasenList = [x.split("\t") for x in chasen.parse(dialogueLog[i][2]).split("\n")]
            chasenTList = pd.DataFrame(chasenList).T.values.tolist()

            # 。！？のどれかが含まれている，かつ，ログの最後に。！？のいずれかが含まれる
            if ((dialogueLog[i][2].find("。") >= 0 or dialogueLog[i][2].find("！") >= 0 or dialogueLog[i][2].find("？") >= 0) and re.search("(。|！|？)$", dialogueLog[i][2]) != None):
                logLoop = wakatiList.count("。") + wakatiList.count("！") + wakatiList.count("？")
            # 。！？のどれかが含まれている，かつ，ログの最後に。！？がない
            elif (dialogueLog[i][2].find("。") >= 0 or dialogueLog[i][2].find("！") >= 0 or dialogueLog[i][2].find("？") >= 0) and re.search("(。|！|？)$", dialogueLog[i][2]) == None:
                logLoop = wakatiList.count("。") + wakatiList.count("！") + wakatiList.count("？") + 1
            # 何もついていない
            else:
                logLoop = 1

            # 文の開始と終了位置を取得
            ls = listSlice(logLoop)[0]
            le = listSlice(logLoop)[1]

            # 発話文毎ループ
            for logL in range(logLoop):
                # 発話文の開始と終了を指定する
                lS = ls[logL]
                lE = le[logL] + 1

                # wakatiとchasenを行い，リスト化
                wakatiList = wakati.parse(dialogueLog[i][2][lS:lE]).split(" ")[:-1]
                chasenList = [x.split("\t") for x in chasen.parse(dialogueLog[i][2][lS:lE]).split("\n")]
                chasenTList = pd.DataFrame(chasenList).T.values.tolist()
                        
                # /startもしくはnullは判定外
                if dialogueLog[i][2][lS:lE] == "/start" or dialogueLog[i][2][lS:lE] == "null":
                    # nowEos.append("")
                    pass
                # P判定（「お気になさらず」）
                elif dialogueLog[i][2][lS:lE].find("お気になさらず。") >= 0 or dialogueLog[i][2][lS:lE].find("お気になさらず！") >= 0  or dialogueLog[i][2][lS:lE].find("お気になさらず？") >= 0 or (dialogueLog[i][2][lS:lE].find("お気になさらず") >= 0 and len(wakatiList) >= 4):
                    nowEos.append("P")
                    userCountP += 1
                # P判定（定型表現）
                elif ((listFind(teikeiList1, dialogueLog[i][2][lS:lE]) == True) or (len(wakatiList) == 1 and "こんにちは" in wakatiList[-1]) or (len(wakatiList) >= 2 and "こんにちは" in wakatiList[-1]) or \
                     (len(wakatiList) == 1 and "ごめんなさい" in wakatiList[-1]) or (len(wakatiList) >= 2 and "ごめんなさい" in wakatiList[-1]) or (len(wakatiList) == 1 and "はじめまして" in wakatiList[-1]) or (len(wakatiList) >= 2 and "はじめまして" in wakatiList[-1]) or \
                     (len(wakatiList) == 1 and "こんばんは" in wakatiList[-1]) or (len(wakatiList) >= 2 and "こんばんは" in wakatiList[-1]) or (listFind(teikeiList2, dialogueLog[i][2][lS:lE]) == True)):
                    nowEos.append("P")
                    userCountP += 1                  
                # P*判定（「です。/ます。」＋α）
                elif re.search("(です((よ|か|ね|けど|から|って)+)。)",  dialogueLog[i][2][lS:lE]) or re.search("(ます((よ|か|ね|けど|から|って)+)。)",  dialogueLog[i][2][lS:lE]):
                    nowEos.append("P*")
                    userCountP1 += 1                   
                # P判定（「です。/ます。」）
                elif dialogueLog[i][2][lS:lE].find("です。") >= 0 or dialogueLog[i][2][lS:lE].find("ます。") >= 0:
                    nowEos.append("P")
                    userCountP += 1                  
                # P*判定（「です/ます」＋α　「です！/ます！」＋α　「です？/ます？」＋α）
                elif re.search("(です((よ|か|ね|けど|から|って)+)$)",  dialogueLog[i][2][lS:lE]) or re.search("(ます((よ|か|ね|けど|から|って)+)$)",  dialogueLog[i][2][lS:lE]) or \
                     re.search("(((です|ます)((よ|か|ね|けど|から|って)+)(！|？))$)",  dialogueLog[i][2][lS:lE]):
                    nowEos.append("P*")
                    userCountP1 += 1
                # P判定（「です/ます」　「です！/ます！」　「です？/ます？」）
                elif re.search("(です|ます)$", dialogueLog[i][2][lS:lE]) or re.search("(です！|ます！|です？|ます？)$", dialogueLog[i][2][lS:lE]):
                    nowEos.append("P")
                    userCountP += 1
                # P*判定（「ください。」＋α）
                elif dialogueLog[i][2][lS:lE].find("くださいね。") >= 0 or dialogueLog[i][2][lS:lE].find("くださいよ。") >= 0:
                    nowEos.append("P*")
                    userCountP1 += 1
                # P判定（「ください。」）
                elif dialogueLog[i][2][lS:lE].find("ください。") >= 0:
                    nowEos.append("P")
                    userCountP += 1
                # P*判定（「ください」＋α　「ください！」＋α　「ください？」＋α）
                elif re.search("(くださいね|くださいね！|くださいね？|くださいよ|くださいよ！|くださいよ？)$", dialogueLog[i][2][lS:lE]):
                    nowEos.append("P*")
                    userCountP1 += 1
                # P判定（「ください」　「ください！」　「ください？」）
                elif re.search("(ください|ください！|ください？)$", dialogueLog[i][2][lS:lE]):
                    nowEos.append("P")
                    userCountP += 1
                # P**判定（「です、/ます、」＋α）
                elif re.search("(です((よ|か|ね|けど|から|って)+)、)",  dialogueLog[i][2][lS:lE]) or re.search("(ます((よ|か|ね|けど|から|って)+)、)",  dialogueLog[i][2][lS:lE]):
                    nowEos.append("P**")
                    userCountP2 += 1
                # P**判定（[「です/ます」＋α（判定できなかった発話用））
                elif re.search("(です((よ|か|ね|けど|から|って)+))",  dialogueLog[i][2][lS:lE]) or re.search("(ます((よ|か|ね|けど|から|って)+))",  dialogueLog[i][2][lS:lE]):
                    nowEos.append("P**")
                    userCountP2 += 1
                # P**判定（「です/ます」（判定できなかった発話用））
                elif dialogueLog[i][2][lS:lE].find("です") >= 0 or dialogueLog[i][2][lS:lE].find("ます") >= 0:
                    nowEos.append("P**")
                    userCountP2 += 1
                # P**判定（「ください」（判定できなかった発話用））
                elif dialogueLog[i][2][lS:lE].find("ください") >= 0:
                    nowEos.append("P**")
                    userCountP2 += 1
                # P*判定（「でした。/ました。」＋α）
                elif re.search("((でしょ|でし|ませ|ましょ|まし|ますれ).([よ|か|ね|けど|から|って]+)。)", dialogueLog[i][2][lS:lE]):
                    nowEos.append("P*")
                    userCountP1 += 1
                # P判定（「でした。/ました。」）
                elif re.search("(((でしょ|でし|ませ|ましょ|まし|ますれ).)。)", dialogueLog[i][2][lS:lE]):
                    nowEos.append("P")
                    userCountP += 1
                # P*判定（「でした/ました」＋α　「でした！/ました！」＋α　「でした？/ました？」＋α）
                elif re.search("((でしょ|でし|ませ|ましょ|まし|ますれ).((よ|か|ね|けど|から|って)+)$)", dialogueLog[i][2][lS:lE]) or \
                     re.search("(((でしょ|でし|ませ|ましょ|まし|ますれ).((よ|か|ね|けど|から|って)+)(！|？))$)", dialogueLog[i][2][lS:lE]):
                    nowEos.append("P*")
                    userCountP1 += 1
                # P判定（「でした/ました」　「でした！/ました！」　「でした？/ました？」）
                elif re.search("(でしょう|でした|ません|ましょう|ました|まして)$", dialogueLog[i][2][lS:lE]) or re.search("((でしょう|でした|ません|ましょう|ました|ますれば)(！|？))$", dialogueLog[i][2][lS:lE]):
                    nowEos.append("P")
                    userCountP += 1
                # P**判定（「でした、/ました、」＋α）
                elif re.search("((でしょ|でし|ませ|ましょ|まし|ますれ).((よ|か|ね|けど|から|って)+)、)", dialogueLog[i][2][lS:lE]):
                    nowEos.append("P**")
                    userCountP2 += 1
                # P**判定（「でした/ました」＋α（判定できなかった発話用））
                elif re.search("((でしょ|でし|ませ|ましょ|まし|ますれ).((よ|か|ね|けど|から|って)+))", dialogueLog[i][2][lS:lE]):
                    nowEos.append("P**")
                    userCountP2 += 1
                # P**判定（「でした/ました」（判定できなかった発話用））
                elif re.search("(でしょう|でした|ません|ましょう|ました|ますれば|まして)", dialogueLog[i][2][lS:lE]):
                    nowEos.append("P**")
                    userCountP2 += 1
                # NM判定（応答　うなずき）
                elif dialogueLog[i][2][lS:lE] in NMlist:
                    nowEos.append("NM")
                    userCountNM += 1
                # NM判定（形容動詞終了）
                elif (len(chasenTList[2]) == 3 and "形容動詞" in chasenTList[3][-3]) or (len(chasenTList[2]) >= 4 and "形容動詞" in chasenTList[3][-3]) or (len(chasenTList[2]) >= 4 and "形容動詞" in chasenTList[3][-4] and "記号" in chasenTList[3][-3]):
                    nowEos.append("NM")
                    userCountNM += 1
                # NM判定（名詞止め（体言止め））
                elif (len (chasenTList[2]) == 3 and "名詞" in chasenTList[3][-3]) or (len (chasenTList[2]) >= 4 and "名詞" in chasenTList[3][-3]) or (len (chasenTList[2]) >= 4 and "名詞" in chasenTList[3][-4] and "記号" in chasenTList[3][-3]):
                    nowEos.append("NM")
                    userCountNM += 1
                # NM判定（言いよどみ）
                elif re.search("…。$", dialogueLog[i][2][lS:lE]) or re.search("・・・。$", dialogueLog[i][2][lS:lE]) or re.search("…$", dialogueLog[i][2][lS:lE]) or re.search("・・・$", dialogueLog[i][2][lS:lE]):
                    nowEos.append("NM")
                    userCountNM += 1
                # NM判定（文末が名詞＋助詞）
                elif ((len(chasenTList[2]) == 4 and ("名詞" in chasenTList[3][-4] and "助詞" in chasenTList[3][-3])) or (len(chasenTList[2]) >= 5 and ("名詞" in chasenTList[3][-5] and "助詞" in chasenTList[3][-4] and "記号" in chasenTList[3][-3])) or \
                     (len(chasenTList[2]) >= 5 and ("名詞" in chasenTList[3][-4] and "助詞" in chasenTList[3][-3]))):
                    nowEos.append("NM")
                    userCountNM += 1
                # N判定
                else:
                    nowEos.append("N")
                    userCountN += 1 

            # セルに書き込み
            ws[i + 1][3].value = '　'.join(nowEos)

            # スピーチレベルシフトの判定
            # 1発話目と3発話目はスピレベの判定は必要なし
            if i != 1 and i != 3:
                # nowEosがNMのみの場合は，シフトなし
                if "P" not in nowEos and "P*" not in nowEos and "P**" not in nowEos and "N" not in nowEos:
                    pass
                else:
                    # lastTimeEosの決定
                    for eosN in range(1, i, 2):
                        eos = ws[i - eosN][3].value
                        # 判定にNM以外が含まれていたらbreak
                        if "P" in eos or re.search("(N　|N$)", eos) is not None:
                            break
                    lastTimeEos = eos.split("　")
                    # lastTimeEosがNMだった場合はシフトなし
                    if "P" not in lastTimeEos and "P*" not in lastTimeEos and "P**" not in lastTimeEos and "N" not in lastTimeEos:
                        pass
                    # lastTimeEosがアップシフトの時
                    elif ws[i - eosN][5].value == "アップシフト":
                        # nowEosに非丁寧体が含まれていたらダウンシフト
                        if "N" in nowEos:
                            ws[i + 1][5].value = "ダウンシフト"
                        # nowEosに非丁寧体が含まれていなかったらシフトなし
                        else:
                            pass
                    # lastTimeEosがダウンシフトの時
                    elif ws[i - eosN][5].value == "ダウンシフト":
                        # nowEosに丁寧体が含まれていたらアップシフト
                        if "P" in nowEos or "P*" in nowEos or "P**" in nowEos:
                            ws[i + 1][5].value = "アップシフト"
                        # nowEosに丁寧体が含まれていなかったらシフトなし
                        else:
                            pass
                    # lastTimeEosでシフトは起きてない時
                    else:
                        # lastTimeEosが丁寧体の時
                        if "P" in lastTimeEos or "P*" in lastTimeEos or "P**" in lastTimeEos:
                            # nowEosが非丁寧体の時
                            if "N" in nowEos:
                                ws[i + 1][5].value = "ダウンシフト"
                            # nowEosが丁寧体の時
                            else:
                                pass                                
                        # lastTimeEosが非丁寧体の時
                        else:
                            # nowEosが丁寧体の時
                            if "P" in nowEos or "P*" in nowEos or "P**" in nowEos:
                                ws[i + 1][5].value = "アップシフト"
                            # nowEosが非丁寧体の時
                            else:
                                pass
			
            # 判定上書き
            lastTimeEos = nowEos
            nowEos = []
        
        # システム発話
        else:
            backColor = systembcColor

            # 文末判定
            if i <= 32:
                with open("./rule/" + str(math.ceil(i/2)).zfill(2) + ".txt", "r", encoding = "utf-8") as f:
                    ruleDictList = [{"sentens":s_line.split("\t")[0], "conditionStr":s_line.split("\t")[1].replace("\n", "").replace(" ", ""), \
                                    "conditionList":s_line.split("\t")[1].replace("\n", "").replace(" ", "").replace("||", "&&").split("&&")} for s_line in f]

				# 各ルールごとの距離をいれる
                levenRuleList = [leven(ruleDict["sentens"].split(",")[0], dialogueLog[i][2]) for ruleDict in ruleDictList]
				
				# 各ルール内の文章ごとの距離を入れる
                levenTextList = [leven(x,dialogueLog[i][2]) for x in ruleDictList[levenRuleList.index(max(levenRuleList))]["sentens"].split(",")]
				
				# 各ルールの各文章内で距離が一番近かった文章からモードの判定を行う
                if levenTextList.index(max(levenTextList)) == 0:
                    ws[i + 1][4].value = "丁寧体"
                    sysCountT += 1
                elif levenTextList.index(max(levenTextList)) == 1:
                    ws[i + 1][4].value = "非丁寧体"
                    sysCountH += 1
                else:
                    ws[i + 1][4].value = "混在"
                    sysCountK += 1

        # テキストの部分は折り返して全体表示にする
        ws[i + 1][2].alignment = Alignment(wrapText = True)
        if len(dialogueLog)-1 != i:
            ruledLine = thindotBorder
        else:
            ruledLine = topdotBorder

        # 背景色と罫線引く
        for j in range(6):
            ws[i + 1][j].fill = backColor
            ws[i + 1][j].border = ruledLine

    # ユーザ丁寧体率
    userTeineitai = (userCountP + userCountP1 + userCountP2) / (userCountP + userCountP1 + userCountP2 + userCountNM + userCountN) * 100

    # システム丁寧体率
    sysTeineitai = sysCountT / (sysCountT + sysCountK + sysCountH) * 100

    # その他書き込み
    ws["H1"].value = "ユーザ年齢"
    ws["H2"].value = "ユーザ丁寧体率（%）"
    ws["H3"].value = "システム丁寧体率（%）"
    ws["H1"].border = thinBorder
    ws["H2"].border = thinBorder
    ws["H3"].border = thinBorder
    ws.column_dimensions["H"].width = 18
    ws["I2"].value = userTeineitai
    ws["I3"].value = sysTeineitai
    ws["I1"].border = thinBorder
    ws["I2"].border = thinBorder
    ws["I3"].border = thinBorder

    ws["H5"].value = "P"
    ws["H6"].value = "P*"
    ws["H7"].value = "P**"
    ws["H8"].value = "NM"
    ws["H9"].value = "N"
    ws["H5"].border = thinBorder
    ws["H6"].border = thinBorder
    ws["H7"].border = thinBorder
    ws["H8"].border = thinBorder
    ws["H9"].border = thinBorder
    ws["I5"].value = userCountP
    ws["I6"].value = userCountP1
    ws["I7"].value = userCountP2
    ws["I8"].value = userCountNM
    ws["I9"].value = userCountN
    ws["I5"].border = thinBorder
    ws["I6"].border = thinBorder
    ws["I7"].border = thinBorder
    ws["I8"].border = thinBorder
    ws["I9"].border = thinBorder

    # ユーザ年齢は元xlsxファイルのセルからコピーして書き込む
    copy = ds.cell(row=1, column=9).value
    ws.cell(row=1, column=9, value=copy)
        
    # 条件ごとのカウント
    if file.find("hiteineitai") >= 0:
        hiteineitaiUserCountP += userCountP
        hiteineitaiUserCountP1 += userCountP1
        hiteineitaiUserCountP2 += userCountP2
        hiteineitaiUserCountNM += userCountNM
        hiteineitaiUserCountN += userCountN
    elif file.find("switching") >= 0:
        switchUserCountP += userCountP
        switchUserCountP1 += userCountP1
        switchUserCountP2 += userCountP2
        switchUserCountNM += userCountNM
        switchUserCountN += userCountN
    else:
        teineitaiUserCountP += userCountP
        teineitaiUserCountP1 += userCountP1
        teineitaiUserCountP2 += userCountP2
        teineitaiUserCountNM += userCountNM
        teineitaiUserCountN += userCountN

    # 保存
    wb.save(file[:-5] + "-label.xlsx")

# 条件ごとの文末カウント
print("丁寧体条件Pカウント" + str(teineitaiUserCountP))
print("丁寧体条件P*カウント" + str(teineitaiUserCountP1))
print("丁寧体条件P**カウント" + str(teineitaiUserCountP2))
print("丁寧体条件NMカウント" + str(teineitaiUserCountNM))
print("丁寧体条件Nカウント" + str(teineitaiUserCountN))
print("文末制御条件Pカウント" + str(switchUserCountP))
print("文末制御条件P*カウント" + str(switchUserCountP1))
print("文末制御条件P**カウント" + str(switchUserCountP2))
print("文末制御条件NMカウント" + str(switchUserCountNM))
print("文末制御条件Nカウント" + str(switchUserCountN))
print("非丁寧体条件Pカウント" + str(hiteineitaiUserCountP))
print("非丁寧体条件P*カウント" + str(hiteineitaiUserCountP1))
print("非丁寧体条件P**カウント" + str(hiteineitaiUserCountP2))
print("非丁寧体条件NMカウント" + str(hiteineitaiUserCountNM))
print("非丁寧体条件Nカウント" + str(hiteineitaiUserCountN))

# 条件ごとの丁寧体率
print("丁寧体条件丁寧体率：" + str(((teineitaiUserCountP + teineitaiUserCountP1 + teineitaiUserCountP2)/(teineitaiUserCountP + teineitaiUserCountP1 + teineitaiUserCountP2 + teineitaiUserCountNM + teineitaiUserCountN)*100)) + "%")
print("文末制御条件丁寧体率：" + str(((switchUserCountP + switchUserCountP1 + switchUserCountP2)/(switchUserCountP + switchUserCountP1 + switchUserCountP2 + switchUserCountNM + switchUserCountN)*100)) + "%")
print("非丁寧体条件丁寧体率：" + str(((hiteineitaiUserCountP + hiteineitaiUserCountP1 + hiteineitaiUserCountP2)/(hiteineitaiUserCountP + hiteineitaiUserCountP1 + hiteineitaiUserCountP2 + hiteineitaiUserCountNM + hiteineitaiUserCountN)*100)) + "%")
